from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse, JsonResponse
from .models import Query
from .serializers import QuerySerializer
from connections.models import Connection
from connections.services import execute_query
import csv
import json
import logging
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)


class QueryViewSet(viewsets.ModelViewSet):
    """ViewSet for SQL queries"""
    serializer_class = QuerySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Query.objects.filter(user=self.request.user).select_related('connection')

    def list(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'status': 'success',
            'data': {'queries': serializer.data}
        })

    def retrieve(self, request, pk=None):
        try:
            query = self.get_queryset().get(pk=pk)
            serializer = self.get_serializer(query)
            return Response({
                'status': 'success',
                'data': {'query': serializer.data}
            })
        except Query.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Query not found'
            }, status=status.HTTP_404_NOT_FOUND)

    def create(self, request):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'status': 'success',
                'data': {'query': serializer.data}
            }, status=status.HTTP_201_CREATED)
        return Response({
            'status': 'error',
            'message': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None):
        try:
            query = self.get_queryset().get(pk=pk)
            serializer = self.get_serializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'status': 'success',
                    'message': 'Query updated successfully'
                })
            return Response({
                'status': 'error',
                'message': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except Query.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Query not found'
            }, status=status.HTTP_404_NOT_FOUND)

    def destroy(self, request, pk=None):
        try:
            query = self.get_queryset().get(pk=pk)
            query.delete()
            return Response({
                'status': 'success',
                'message': 'Query deleted successfully'
            })
        except Query.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Query not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def execute(self, request, pk=None):
        """Execute the SQL query"""
        from .sql_validator import validate_sql_query, sanitize_error_message

        try:
            query = self.get_queryset().get(pk=pk)

            # Security: Validate saved query before execution
            user_is_admin = request.user.is_staff or request.user.is_superuser
            is_valid, error_message = validate_sql_query(query.sql, user_is_admin=user_is_admin)

            if not is_valid:
                logger.warning(
                    f"Saved query {pk} validation failed for user {request.user.id}: {error_message}"
                )
                return Response({
                    'status': 'error',
                    'message': f"Query validation failed: {error_message}"
                }, status=status.HTTP_400_BAD_REQUEST)

            result = execute_query(query.connection, query.sql)
            return Response({
                'status': 'success',
                'data': result
            })
        except Query.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Query not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Query execution error: {str(e)}")
            # Security: Sanitize error messages
            safe_message = sanitize_error_message(str(e))
            return Response({
                'status': 'error',
                'message': safe_message
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def execute_raw(self, request):
        """Execute raw SQL query"""
        from .sql_validator import validate_sql_query, sanitize_error_message

        try:
            connection_id = request.data.get('connection_id')
            sql = request.data.get('sql')

            if not connection_id or not sql:
                return Response({
                    'status': 'error',
                    'message': 'connection_id and sql are required'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Security: Validate SQL query before execution
            user_is_admin = request.user.is_staff or request.user.is_superuser
            is_valid, error_message = validate_sql_query(sql, user_is_admin=user_is_admin)

            if not is_valid:
                logger.warning(
                    f"SQL validation failed for user {request.user.id}: {error_message}"
                )
                return Response({
                    'status': 'error',
                    'message': error_message
                }, status=status.HTTP_400_BAD_REQUEST)

            connection = Connection.objects.get(pk=connection_id, user=request.user)
            result = execute_query(connection, sql)
            return Response({
                'status': 'success',
                'data': result
            })
        except Connection.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Connection not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Raw query execution error: {str(e)}")
            # Security: Sanitize error messages to avoid exposing sensitive info
            safe_message = sanitize_error_message(str(e))
            return Response({
                'status': 'error',
                'message': safe_message
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def export_csv(self, request):
        """Export query results to CSV"""
        try:
            connection_id = request.data.get('connection_id')
            sql = request.data.get('sql')
            filename = request.data.get('filename', 'export.csv')

            if not connection_id or not sql:
                return Response({
                    'status': 'error',
                    'message': 'connection_id and sql are required'
                }, status=status.HTTP_400_BAD_REQUEST)

            connection = Connection.objects.get(pk=connection_id, user=request.user)
            result = execute_query(connection, sql, max_rows=50000)  # Higher limit for exports

            # Create CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            writer = csv.writer(response)

            # Write headers
            if result['columns']:
                writer.writerow([col['name'] for col in result['columns']])

            # Write data
            for row in result['rows']:
                writer.writerow([row.get(col['name'], '') for col in result['columns']])

            logger.info(f"Exported {len(result['rows'])} rows to CSV")
            return response

        except Connection.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Connection not found'
            }, status=404)
        except Exception as e:
            logger.error(f"CSV export error: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

    @action(detail=False, methods=['post'])
    def export_excel(self, request):
        """Export query results to Excel"""
        try:
            connection_id = request.data.get('connection_id')
            sql = request.data.get('sql')
            filename = request.data.get('filename', 'export.xlsx')

            if not connection_id or not sql:
                return Response({
                    'status': 'error',
                    'message': 'connection_id and sql are required'
                }, status=status.HTTP_400_BAD_REQUEST)

            connection = Connection.objects.get(pk=connection_id, user=request.user)
            result = execute_query(connection, sql, max_rows=50000)

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Query Results"

            # Style for headers
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Write headers
            if result['columns']:
                for col_idx, col in enumerate(result['columns'], 1):
                    cell = ws.cell(row=1, column=col_idx, value=col['name'])
                    cell.font = header_font
                    cell.alignment = header_alignment

            # Write data
            for row_idx, row in enumerate(result['rows'], 2):
                for col_idx, col in enumerate(result['columns'], 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col['name'], ''))

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            logger.info(f"Exported {len(result['rows'])} rows to Excel")
            return response

        except Connection.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Connection not found'
            }, status=404)
        except Exception as e:
            logger.error(f"Excel export error: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

    @action(detail=False, methods=['post'])
    def export_json(self, request):
        """Export query results to JSON"""
        try:
            connection_id = request.data.get('connection_id')
            sql = request.data.get('sql')
            filename = request.data.get('filename', 'export.json')

            if not connection_id or not sql:
                return Response({
                    'status': 'error',
                    'message': 'connection_id and sql are required'
                }, status=status.HTTP_400_BAD_REQUEST)

            connection = Connection.objects.get(pk=connection_id, user=request.user)
            result = execute_query(connection, sql, max_rows=50000)

            # Create JSON response
            response = HttpResponse(
                json.dumps(result['rows'], indent=2),
                content_type='application/json'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            logger.info(f"Exported {len(result['rows'])} rows to JSON")
            return response

        except Connection.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Connection not found'
            }, status=404)
        except Exception as e:
            logger.error(f"JSON export error: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
